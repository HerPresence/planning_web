import csv
import io
import json
import os
import subprocess
import sys
import tempfile
import traceback
from typing import Optional

from fastapi import APIRouter, File, Form, HTTPException, UploadFile

from db import get_connection

try:
    import gspread
    from google.oauth2.service_account import Credentials
    _GOOGLE_AVAILABLE = True
except ImportError:
    _GOOGLE_AVAILABLE = False

try:
    import openpyxl
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False

try:
    import pyodbc
    _PYODBC_AVAILABLE = True
except ImportError:
    _PYODBC_AVAILABLE = False

router = APIRouter(prefix="/api/pnl-import")

GOOGLE_CREDENTIALS_PATH = os.getenv(
    "GOOGLE_CREDENTIALS_PATH",
    r"T:\planning_web\google_credentials.json",
)

GOOGLE_SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets.readonly",
    "https://www.googleapis.com/auth/drive.readonly",
]


# ─── ensure tables ────────────────────────────────────────────────────────────

def ensure_article_mapping_table():
    conn = get_connection()
    cur = conn.cursor()
    try:
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS article_mapping (
                mapping_id          SERIAL PRIMARY KEY,
                source_id           INTEGER,
                source_system       TEXT NOT NULL DEFAULT '',
                source_article_id   TEXT,
                source_article_name TEXT,
                article_id          INTEGER,
                comment             TEXT,
                is_active           BOOLEAN DEFAULT TRUE
            )
            """
        )
        conn.commit()

        cur.execute(
            """
            SELECT 1 FROM information_schema.columns
            WHERE table_name = 'article_mapping' AND column_name = 'source_id'
            """
        )
        if not cur.fetchone():
            cur.execute("SET lock_timeout = '3s'")
            cur.execute("ALTER TABLE article_mapping ADD COLUMN source_id INTEGER")
            conn.commit()

        # Ensure source_system has a default so old NOT NULL rows don't block new inserts
        try:
            cur.execute("SET lock_timeout = '3s'")
            cur.execute(
                "ALTER TABLE article_mapping "
                "ALTER COLUMN source_system SET DEFAULT ''"
            )
            conn.commit()
        except Exception:
            conn.rollback()

    except Exception as exc:
        print(f"[startup] ensure_article_mapping_table warning: {exc}")
        try:
            conn.rollback()
        except Exception:
            pass
    finally:
        cur.close()
        conn.close()


def ensure_department_mapping_table():
    conn = get_connection()
    cur = conn.cursor()
    try:
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS department_import_mapping (
                id                       SERIAL PRIMARY KEY,
                source_id                INTEGER NOT NULL,
                external_department_code TEXT,
                external_department_name TEXT,
                internal_department_id   TEXT NOT NULL,
                is_active                BOOLEAN DEFAULT TRUE
            )
            """
        )
        conn.commit()

        # Migrate existing INTEGER column to TEXT if needed
        cur.execute(
            """
            SELECT data_type FROM information_schema.columns
            WHERE table_name = 'department_import_mapping'
              AND column_name = 'internal_department_id'
            """
        )
        row = cur.fetchone()
        if row and row[0].lower() in ("integer", "bigint", "smallint"):
            cur.execute(
                "ALTER TABLE department_import_mapping "
                "ALTER COLUMN internal_department_id TYPE TEXT USING internal_department_id::TEXT"
            )
            conn.commit()

    except Exception as exc:
        print(f"[startup] ensure_department_mapping_table warning: {exc}")
        try:
            conn.rollback()
        except Exception:
            pass
    finally:
        cur.close()
        conn.close()


def ensure_pnl_column_mapping_table():
    conn = get_connection()
    cur = conn.cursor()
    try:
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS pnl_import_column_mapping (
                id              SERIAL PRIMARY KEY,
                source_id       INTEGER NOT NULL UNIQUE,
                period_col      TEXT DEFAULT '',
                dept_code_col   TEXT DEFAULT '',
                dept_name_col   TEXT DEFAULT '',
                article_code_col TEXT DEFAULT '',
                article_name_col TEXT DEFAULT '',
                amount_col      TEXT DEFAULT '',
                comment_col     TEXT DEFAULT '',
                updated_at      TIMESTAMP DEFAULT NOW()
            )
            """
        )
        conn.commit()
    except Exception as exc:
        print(f"[startup] ensure_pnl_column_mapping_table warning: {exc}")
        try:
            conn.rollback()
        except Exception:
            pass
    finally:
        cur.close()
        conn.close()


# ─── source reading helpers ───────────────────────────────────────────────────

def _read_google_sheet(sheet_url: str, sheet_name: Optional[str]) -> tuple:
    """Returns (rows: list[dict], sheet_names: list[str])"""
    if not _GOOGLE_AVAILABLE:
        raise RuntimeError(
            "gspread / google-auth не встановлено. Встановіть: pip install gspread google-auth"
        )
    if not os.path.exists(GOOGLE_CREDENTIALS_PATH):
        raise FileNotFoundError(
            f"Google credentials не знайдено: {GOOGLE_CREDENTIALS_PATH}"
        )
    creds = Credentials.from_service_account_file(
        GOOGLE_CREDENTIALS_PATH, scopes=GOOGLE_SCOPES
    )
    gc = gspread.authorize(creds)
    spreadsheet = gc.open_by_url(sheet_url)
    sheet_names = [ws.title for ws in spreadsheet.worksheets()]
    ws = spreadsheet.worksheet(sheet_name) if sheet_name else spreadsheet.sheet1
    return ws.get_all_records(), sheet_names


def _read_file(file_bytes: bytes, filename: str, sheet_name: Optional[str] = None) -> tuple:
    """Returns (rows: list[dict], sheet_names: list[str])"""
    fname = filename.lower()
    if fname.endswith(".csv"):
        content = file_bytes.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(content))
        return [dict(row) for row in reader], []
    if fname.endswith((".xlsx", ".xls")):
        if not _OPENPYXL_AVAILABLE:
            raise RuntimeError(
                "openpyxl не встановлено. Встановіть: pip install openpyxl"
            )
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        available_sheets = wb.sheetnames
        if sheet_name and sheet_name in available_sheets:
            ws = wb[sheet_name]
        else:
            ws = wb.active
        headers = [
            str(c.value).strip() if c.value is not None else f"col_{i}"
            for i, c in enumerate(ws[1])
        ]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(v is None for v in row):
                continue
            rows.append(
                {headers[i]: ("" if row[i] is None else row[i]) for i in range(len(headers))}
            )
        return rows, available_sheets
    raise ValueError(f"Непідтримуваний формат файлу: {filename}")


# ── Constants ─────────────────────────────────────────────────────────────────

# SQL Server ODBC drivers (for sql_odbc mode only)
_ODBC_DRIVERS_MSSQL = [
    "ODBC Driver 18 for SQL Server",
    "ODBC Driver 17 for SQL Server",
    "ODBC Driver 13 for SQL Server",
    "SQL Server Native Client 11.0",
    "SQL Server",
]

# PowerShell script that executes DAX via ADOMD.NET (OLE DB — no Python .NET binding needed)
_PS_SSAS_SCRIPT = os.path.normpath(
    os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "tools", "read_ssas_dax.ps1")
)

_SSAS_ADOMD_NOT_FOUND_MSG = (
    "Microsoft Analysis Services ADOMD.NET client не знайдено.\n"
    "Встановіть Microsoft OLE DB Provider for Analysis Services:\n"
    "  https://learn.microsoft.com/en-us/analysis-services/client-libraries\n"
    "  або запустіть install_olap.bat"
)


# ── DB config helper ──────────────────────────────────────────────────────────

def _get_olap_source_config(source_id: int):
    """Fetch OLAP source settings from import_sources."""
    conn_db = get_connection()
    cur = conn_db.cursor()
    try:
        cur.execute(
            """
            SELECT db_server, db_port, db_database, db_cube_model,
                   db_login, db_password, db_query
            FROM import_sources
            WHERE id = %s AND is_active = TRUE
            """,
            (source_id,),
        )
        return cur.fetchone()
    finally:
        cur.close()
        conn_db.close()


# ── Connection string builders ────────────────────────────────────────────────

def _build_mssql_conn_str(db_server, db_port, db_database, db_login, db_password,
                           driver) -> str:
    """ODBC connection string for regular SQL Server (pyodbc)."""
    server_part = f"{db_server},{db_port}" if db_port else db_server
    db_part = f";DATABASE={db_database}" if db_database else ""
    if db_login and db_password:
        return (
            f"DRIVER={{{driver}}};"
            f"SERVER={server_part}{db_part};"
            f"UID={db_login};PWD={db_password};"
            f"TrustServerCertificate=yes"
        )
    return (
        f"DRIVER={{{driver}}};"
        f"SERVER={server_part}{db_part};"
        f"Trusted_Connection=yes;TrustServerCertificate=yes"
    )


# ── SSAS Tabular / DAX via PowerShell + ADOMD.NET ────────────────────────────

def _run_ssas_ps(db_server, db_port, db_database, db_login, db_password,
                 db_query, max_rows: int = 0, source_label: str = "") -> dict:
    """Call read_ssas_dax.ps1 via subprocess; return parsed JSON dict.

    The DAX query is written to a temp file and passed via -QueryFile to avoid:
    - Windows command-line length limits (32K chars)
    - Newline/encoding issues in quoted arguments that cause PowerShell to
      crash before writing output, producing WinError 233 on the Python side.
    """
    if not os.path.isfile(_PS_SSAS_SCRIPT):
        raise FileNotFoundError(
            f"PowerShell SSAS script not found: {_PS_SSAS_SCRIPT}. "
            "Ensure tools/read_ssas_dax.ps1 exists in the project root."
        )

    server = f"{db_server}:{db_port}" if db_port else db_server

    # All three temp files are tracked here so the finally block can delete them.
    query_file  = None
    stdout_file = None
    stderr_file = None
    try:
        # ---- STEP 1: create temp files ----
        with tempfile.NamedTemporaryFile(
            mode="w", suffix=".dax", delete=False, encoding="utf-8"
        ) as tf:
            tf.write(db_query)
            query_file = tf.name
        # stdout/stderr written to files instead of pipes — eliminates WinError 233
        # which occurs when subprocess.run uses PIPE handles that the Windows process
        # manager invalidates when the parent has no console (uvicorn runs headless).
        with tempfile.NamedTemporaryFile(suffix=".out", delete=False) as tf:
            stdout_file = tf.name
        with tempfile.NamedTemporaryFile(suffix=".err", delete=False) as tf:
            stderr_file = tf.name
        print(
            f"[SSAS-DAX] STEP1 query_file={query_file!r}  "
            f"exists={os.path.isfile(query_file)}  size={os.path.getsize(query_file)}  "
            f"stdout_file={stdout_file!r}  stderr_file={stderr_file!r}"
        )

        # ---- STEP 2: build command ----
        cmd = [
            "powershell.exe",
            "-NoProfile", "-NonInteractive",
            "-ExecutionPolicy", "Bypass",
            "-File", _PS_SSAS_SCRIPT,
            "-Server", server,
            "-Database", db_database,
            "-QueryFile", query_file,
        ]
        if db_login:
            cmd += ["-Login", db_login]
        if db_password:
            cmd += ["-Password", db_password]
        if max_rows > 0:
            cmd += ["-MaxRows", str(max_rows)]

        print(f"[SSAS-DAX] source={source_label!r}  server={server!r}  catalog={db_database!r}")
        print(f"[SSAS-DAX] query ({len(db_query)} chars): {db_query[:200]}...")
        print(f"[SSAS-DAX] STEP2 ps_script={_PS_SSAS_SCRIPT!r}  script_exists={os.path.isfile(_PS_SSAS_SCRIPT)}")
        print(f"[SSAS-DAX] STEP2 cmd={cmd}")
        print(f"[SSAS-DAX] STEP2 platform={sys.platform}")

        # ---- STEP 3: run process, output to temp files (no pipes) ----
        create_flags = subprocess.CREATE_NO_WINDOW if sys.platform == "win32" else 0
        print(f"[SSAS-DAX] STEP3 starting Popen with file output  platform={sys.platform}")
        try:
            with open(stdout_file, "wb") as out_fh, open(stderr_file, "wb") as err_fh:
                proc = subprocess.Popen(cmd, stdout=out_fh, stderr=err_fh, creationflags=create_flags)
            print(f"[SSAS-DAX] STEP3 Popen started  pid={proc.pid}")
            try:
                returncode = proc.wait(timeout=130)
            except subprocess.TimeoutExpired:
                proc.kill()
                proc.wait()
                raise RuntimeError("PowerShell SSAS script перевищив таймаут (130 с).")
        except FileNotFoundError:
            raise RuntimeError(
                "powershell.exe не знайдено. Переконайтесь, що backend запущено на Windows."
            )
        except RuntimeError:
            raise
        except OSError as exc:
            tb = traceback.format_exc()
            raise RuntimeError(
                f"[SSAS-DAX] STEP3 FAILED Popen OSError "
                f"(WinError {exc.winerror if hasattr(exc, 'winerror') else '?'}): {exc}\n"
                f"source={source_label!r}  server={server!r}  catalog={db_database!r}\n"
                f"traceback:\n{tb}"
            )
        except Exception as exc:
            tb = traceback.format_exc()
            raise RuntimeError(
                f"[SSAS-DAX] STEP3 FAILED Popen unexpected {type(exc).__name__}: {exc}\n"
                f"traceback:\n{tb}"
            )
        print(f"[SSAS-DAX] STEP3 process done  returncode={returncode}")

        # ---- STEP 4: read output files and parse ----
        with open(stdout_file, "r", encoding="utf-8", errors="replace") as f:
            stdout = f.read().strip()
        with open(stderr_file, "r", encoding="utf-8", errors="replace") as f:
            stderr = f.read().strip()

        print(f"[SSAS-DAX] STEP4 exit={returncode}  stdout_len={len(stdout)}  stderr_len={len(stderr)}")
        if stderr:
            print(f"[SSAS-DAX] STEP4 stderr: {stderr[:800]}")
        if stdout:
            print(f"[SSAS-DAX] STEP4 stdout first 400: {stdout[:400]}")

        if not stdout:
            raise RuntimeError(
                f"PowerShell не повернув жодних даних (exit {returncode}).\n"
                f"source={source_label!r}  server={server!r}  catalog={db_database!r}\n"
                f"query_chars={len(db_query)}  query_file_used={query_file!r}\n"
                f"stderr: {stderr[:600]}"
            )

        json_line = ""
        for line in reversed(stdout.splitlines()):
            line = line.strip()
            if line.startswith("{"):
                json_line = line
                break

        if not json_line:
            raise RuntimeError(
                f"PowerShell не повернув JSON (exit {returncode}).\n"
                f"source={source_label!r}  server={server!r}\n"
                f"stdout (last 600): {stdout[-600:]}"
            )

        try:
            data = json.loads(json_line)
        except json.JSONDecodeError as exc:
            raise RuntimeError(
                f"PowerShell повернув невалідний JSON: {exc}\n"
                f"source={source_label!r}\n"
                f"raw: {json_line[:400]}"
            )

        return data

    finally:
        for _path in (query_file, stdout_file, stderr_file):
            if _path:
                try:
                    os.unlink(_path)
                except OSError:
                    pass
        print("[SSAS-DAX] temp files cleaned up")


def _read_ssas_dax(source_id: int) -> tuple:
    """Execute a DAX query against SSAS Tabular via PowerShell + ADOMD.NET."""
    cfg = _get_olap_source_config(source_id)
    if not cfg:
        raise ValueError(f"Джерело ID={source_id} не знайдено або неактивне")

    db_server, db_port, db_database, db_cube_model, db_login, db_password, db_query = cfg

    # Fetch display name for logging (separate query to avoid changing cfg tuple)
    conn_n = get_connection(); cur_n = conn_n.cursor()
    try:
        cur_n.execute("SELECT source_name FROM import_sources WHERE id = %s", (source_id,))
        row_n = cur_n.fetchone()
    finally:
        cur_n.close(); conn_n.close()
    source_name  = (row_n[0] if row_n else "") or ""
    source_label = f"id={source_id} name={source_name!r}"

    print(f"[SSAS-DAX] _read_ssas_dax: {source_label}")
    print(f"[SSAS-DAX] server={db_server!r}  database={db_database!r}")

    if not db_server:
        raise ValueError(f"[{source_label}] db_server не вказано у налаштуваннях джерела")
    if not db_database:
        raise ValueError(f"[{source_label}] db_database (Initial Catalog) не вказано у налаштуваннях джерела")
    if not db_query:
        raise ValueError(f"[{source_label}] db_query (DAX) не вказано у налаштуваннях джерела")

    print(f"[SSAS-DAX] query ({len(db_query)} chars): {db_query[:200]}...")

    try:
        data = _run_ssas_ps(
            db_server, db_port, db_database, db_login, db_password,
            db_query, source_label=source_label,
        )
    except Exception as exc:
        tb = traceback.format_exc()
        print(f"[SSAS-DAX] FAILED for {source_label}:\n{tb}")
        raise

    if "error" in data:
        raise RuntimeError(
            f"SSAS DAX помилка [{source_label}] ({db_server}/{db_database}): {data['error']}\n"
            f"Перевірте: сервер доступний (порт 2383/2382), каталог '{db_database}' існує в SSAS."
        )

    columns = data.get("columns", [])
    rows = data.get("rows", [])
    print(f"[SSAS-DAX] success: {len(rows)} рядків, колонки: {columns}")
    return rows, []


# ── Regular SQL Server via ODBC (pyodbc) ──────────────────────────────────────

def _read_sql_odbc(source_id: int) -> tuple:
    """Execute a SQL query against SQL Server via ODBC Driver (pyodbc)."""
    if not _PYODBC_AVAILABLE:
        raise RuntimeError(
            "pyodbc не встановлено. Запустіть: pip install pyodbc"
        )

    cfg = _get_olap_source_config(source_id)
    if not cfg:
        raise ValueError(f"Джерело ID={source_id} не знайдено або неактивне")

    db_server, db_port, db_database, db_cube_model, db_login, db_password, db_query = cfg

    if not db_server:
        raise ValueError("db_server не вказано у налаштуваннях джерела")
    if not db_query:
        raise ValueError("db_query не вказано у налаштуваннях джерела")

    print(f"[SQL-ODBC] server={db_server!r}  db={db_database!r}")
    errors = []

    for driver in _ODBC_DRIVERS_MSSQL:
        try:
            conn_str = _build_mssql_conn_str(
                db_server, db_port, db_database, db_login, db_password, driver
            )
            print(f"[SQL-ODBC] Спроба: {driver!r}")
            conn = pyodbc.connect(conn_str, timeout=30)
            print(f"[SQL-ODBC] {driver!r} — підключено")
            cur = conn.cursor()
            cur.execute(db_query)
            columns = [desc[0] for desc in cur.description]
            rows = [
                {columns[i]: ("" if r[i] is None else r[i]) for i in range(len(columns))}
                for r in cur.fetchall()
            ]
            conn.close()
            print(f"[SQL-ODBC] {len(rows)} рядків, колонки: {columns}")
            return rows, []
        except Exception as exc:
            msg = f"{driver}: {exc}"
            print(f"[SQL-ODBC] {msg}")
            errors.append(msg)

    raise RuntimeError(
        f"Не вдалося підключитись до SQL Server ({db_server}/{db_database}). "
        f"Встановіть 'ODBC Driver 17 for SQL Server' або новіший. "
        f"Деталі: {'; '.join(errors)}"
    )


# ── Backward-compat wrapper (olap_sql legacy) ─────────────────────────────────

def _read_olap(source_id: int) -> tuple:
    """Legacy olap_sql alias — routes to SSAS DAX via PowerShell."""
    return _read_ssas_dax(source_id)


def _normalize_code(val) -> str:
    """Convert Excel numeric codes to string: 2.0 → '2', 1014.0 → '1014'."""
    if val is None:
        return ""
    if isinstance(val, float) and val.is_integer():
        return str(int(val))
    return str(val).strip()


# ─── lookup helpers ───────────────────────────────────────────────────────────

def _lookup_department(cur, source_id: int, ext_code: str, ext_name: str):
    if ext_code:
        cur.execute(
            """
            SELECT internal_department_id FROM department_import_mapping
            WHERE source_id = %s AND external_department_code = %s AND is_active = TRUE
            """,
            (source_id, str(ext_code)),
        )
        found = cur.fetchall()
        if len(found) == 1:
            return found[0][0], None
        if len(found) > 1:
            return None, "department_ambiguous_mapping"

    if ext_name:
        cur.execute(
            """
            SELECT internal_department_id FROM department_import_mapping
            WHERE source_id = %s AND external_department_name = %s AND is_active = TRUE
            """,
            (source_id, str(ext_name)),
        )
        found = cur.fetchall()
        if len(found) == 1:
            return found[0][0], None
        if len(found) > 1:
            return None, "department_ambiguous_mapping"

    return None, "department_not_mapped"


def _lookup_article(cur, source_id: int, ext_code: str, ext_name: str):
    if ext_code:
        cur.execute(
            """
            SELECT article_id FROM article_mapping
            WHERE source_id = %s AND source_article_id = %s AND is_active = TRUE
            """,
            (source_id, str(ext_code)),
        )
        found = cur.fetchall()
        if len(found) == 1:
            return found[0][0], None
        if len(found) > 1:
            return None, "article_ambiguous_mapping"

    if ext_name:
        cur.execute(
            """
            SELECT article_id FROM article_mapping
            WHERE source_id = %s AND source_article_name = %s AND is_active = TRUE
            """,
            (source_id, str(ext_name)),
        )
        found = cur.fetchall()
        if len(found) == 1:
            return found[0][0], None
        if len(found) > 1:
            return None, "article_ambiguous_mapping"

    return None, "article_not_mapped"


# ─── shared row processor ─────────────────────────────────────────────────────

def _process_rows(cur, rows, source_id, dept_code_col, dept_name_col,
                  article_code_col, article_name_col, period_col, amount_col, comment_col):
    """Validate every row; return (candidates, skipped, errors).
    candidates is a list of dicts ready to INSERT — no DB writes happen here.
    """
    candidates = []
    skipped = 0
    errors = []

    for row_num, row in enumerate(rows, start=2):
        ext_dept_code = _normalize_code(row.get(dept_code_col)) if dept_code_col else ""
        ext_dept_name = _normalize_code(row.get(dept_name_col)) if dept_name_col else ""
        ext_art_code  = _normalize_code(row.get(article_code_col)) if article_code_col else ""
        ext_art_name  = _normalize_code(row.get(article_name_col)) if article_name_col else ""

        period = _normalize_code(row.get(period_col))
        if not period:
            skipped += 1
            errors.append({"row": row_num, "type": "missing_period", "value": ""})
            continue

        raw_amount = row.get(amount_col, "")
        try:
            amount = float(str(raw_amount).replace(",", ".").strip())
        except (ValueError, TypeError):
            skipped += 1
            errors.append({"row": row_num, "type": "invalid_amount", "value": str(raw_amount)})
            continue

        dept_id, dept_err = _lookup_department(cur, source_id, ext_dept_code, ext_dept_name)
        if dept_err:
            skipped += 1
            errors.append({
                "row": row_num, "type": dept_err,
                "value": ext_dept_code or ext_dept_name,
                "ext_code": ext_dept_code, "ext_name": ext_dept_name,
            })
            continue

        art_id, art_err = _lookup_article(cur, source_id, ext_art_code, ext_art_name)
        if art_err:
            skipped += 1
            errors.append({
                "row": row_num, "type": art_err,
                "value": ext_art_code or ext_art_name,
                "ext_code": ext_art_code, "ext_name": ext_art_name,
            })
            continue

        cur.execute(
            "SELECT holding_name, organization_name, region_name, branch_name, department_name "
            "FROM dim_department WHERE department_id = %s",
            (dept_id,),
        )
        dept_row = cur.fetchone()
        if not dept_row:
            skipped += 1
            errors.append({"row": row_num, "type": "department_not_found_in_dim", "value": str(dept_id)})
            continue

        holding_name, organization_name, region_name, branch_name, department_name = dept_row

        cur.execute(
            "SELECT article_name, pnl_id FROM dim_article WHERE article_id = %s",
            (art_id,),
        )
        art_row = cur.fetchone()
        if not art_row:
            skipped += 1
            errors.append({"row": row_num, "type": "article_not_found_in_dim", "value": str(art_id)})
            continue

        article_name, pnl_id = art_row
        comment = str(row.get(comment_col, "")).strip() if comment_col else ""

        candidates.append({
            "period": period,
            "holding_name": holding_name,
            "organization_name": organization_name,
            "region_name": region_name,
            "branch_name": branch_name,
            "dept_id": str(dept_id),
            "department_name": department_name,
            "art_id": str(art_id),
            "article_name": article_name,
            "pnl_id": str(pnl_id) if pnl_id else "",
            "amount": amount,
            "comment": comment,
        })

    return candidates, skipped, errors


def _count_existing(cur, import_type: str, scenario: str, version_name: str, periods: list) -> int:
    if not periods:
        return 0
    if import_type == "plan":
        cur.execute(
            "SELECT COUNT(*) FROM plan_pnl "
            "WHERE scenario = %s AND version_name = %s AND period = ANY(%s::date[])",
            (scenario, version_name, periods),
        )
    else:
        cur.execute("SELECT COUNT(*) FROM fact_pnl WHERE period = ANY(%s::date[])", (periods,))
    row = cur.fetchone()
    return row[0] if row else 0


# ─── OLAP test connection ─────────────────────────────────────────────────────

@router.get("/test-connection/{source_id}")
def test_olap_connection(source_id: int, mode: str = "auto"):
    """Diagnose DB/OLAP connectivity.

    mode:
      auto          — detect from source_type in import_sources
      olap_ssas_dax — SSAS Tabular via ADOMD.NET/pyadomd (DAX queries)
      olap_sql      — same as olap_ssas_dax (legacy alias)
      sql_odbc      — SQL Server via ODBC Driver 17/18
    """
    result = {
        "source_id": source_id,
        "mode": mode,
        # SSAS/DAX transport — PowerShell + ADOMD.NET
        "ssas_transport": "PowerShell + ADOMD.NET (OLE DB)",
        "ps_script": _PS_SSAS_SCRIPT,
        "ps_script_exists": os.path.isfile(_PS_SSAS_SCRIPT),
        # pyodbc / SQL Server ODBC
        "pyodbc_available": _PYODBC_AVAILABLE,
        "mssql_drivers_found": [],
        # Connection test
        "connection_ok": False,
        "driver_used": None,
        "columns": [],
        "rows_preview": [],
        "attempts": [],
        "server": None,
        "database": None,
        "cube_model": None,
        "query_preview": "",
        "error": None,
    }

    # 1. Source config
    cfg = _get_olap_source_config(source_id)
    if not cfg:
        result["error"] = f"Джерело ID={source_id} не знайдено або неактивне"
        return result

    db_server, db_port, db_database, db_cube_model, db_login, db_password, db_query = cfg
    result["server"]        = f"{db_server}:{db_port}" if db_port else db_server
    result["database"]      = db_database
    result["cube_model"]    = db_cube_model
    result["query_preview"] = (db_query or "")[:300]

    if not db_server:
        result["error"] = "db_server не вказано у налаштуваннях джерела"
        return result
    if not db_query:
        result["error"] = "db_query не вказано у налаштуваннях джерела"
        return result

    # 2. Detect effective mode from source_type when mode == "auto"
    effective_mode = mode
    if effective_mode == "auto":
        conn_db = get_connection()
        cur_db  = conn_db.cursor()
        try:
            cur_db.execute("SELECT source_type FROM import_sources WHERE id = %s", (source_id,))
            row_st = cur_db.fetchone()
            effective_mode = (row_st[0] or "olap_ssas_dax") if row_st else "olap_ssas_dax"
        finally:
            cur_db.close()
            conn_db.close()
    result["mode"] = effective_mode

    # 3. SSAS Tabular / DAX — via PowerShell + ADOMD.NET
    if effective_mode in ("olap_ssas_dax", "olap_sql"):
        label = "PowerShell/ADOMD.NET"
        if not result["ps_script_exists"]:
            result["error"] = (
                f"PowerShell SSAS script not found: {_PS_SSAS_SCRIPT}. "
                "Переконайтесь, що tools/read_ssas_dax.ps1 є в директорії проекту."
            )
            result["attempts"].append({"driver": label, "ok": False, "error": result["error"]})
            return result

        try:
            data = _run_ssas_ps(
                db_server, db_port, db_database, db_login, db_password,
                db_query, max_rows=3
            )
            if "error" in data:
                raise RuntimeError(data["error"])
            cols = data.get("columns", [])
            rows_preview = data.get("rows", [])
            result["connection_ok"] = True
            result["driver_used"]   = label
            result["columns"]       = cols
            result["rows_preview"]  = rows_preview[:3]
            result["attempts"].append({"driver": label, "ok": True, "error": None})
        except Exception as exc:
            err = str(exc)
            result["attempts"].append({"driver": label, "ok": False, "error": err})
            result["error"] = (
                f"SSAS підключення невдале: {err}\n"
                f"Перевірте: сервер {db_server} доступний (порт 2383), "
                f"каталог '{db_database}' існує в SSAS, облікові дані коректні."
            )
        return result

    # 4. SQL Server via ODBC (pyodbc)
    if effective_mode == "sql_odbc":
        if not _PYODBC_AVAILABLE:
            result["error"] = "pyodbc не встановлено. Запустіть: pip install pyodbc"
            return result

        all_drivers = pyodbc.drivers()
        result["mssql_drivers_found"] = [
            d for d in all_drivers if any(k in d for k in ("SQL Server", "MSSQL"))
        ]

        for driver in _ODBC_DRIVERS_MSSQL:
            try:
                conn_str = _build_mssql_conn_str(
                    db_server, db_port, db_database, db_login, db_password, driver
                )
                c = pyodbc.connect(conn_str, timeout=10)
                cur = c.cursor()
                cur.execute(db_query)
                cols = [d[0] for d in cur.description]
                raw = cur.fetchmany(3)
                c.close()
                rows_preview = [
                    {cols[i]: ("" if r[i] is None else str(r[i])) for i in range(len(cols))}
                    for r in raw
                ]
                result["connection_ok"] = True
                result["driver_used"]   = driver
                result["columns"]       = cols
                result["rows_preview"]  = rows_preview
                result["attempts"].append({"driver": driver, "ok": True, "error": None})
                return result
            except Exception as exc:
                result["attempts"].append({"driver": driver, "ok": False, "error": str(exc)})

        result["error"] = "Всі ODBC-драйвери невдалі. Дивіться поле attempts."
        return result

    result["error"] = f"Невідомий mode: {effective_mode!r}. Очікується olap_ssas_dax / sql_odbc."
    return result


# ─── article mapping CRUD ─────────────────────────────────────────────────────

@router.get("/article-mapping")
def get_article_mappings(source_id: Optional[int] = None):
    conn = get_connection()
    cur = conn.cursor()

    if source_id is not None:
        cur.execute(
            """
            SELECT mapping_id, source_id, source_article_id, source_article_name,
                   article_id, comment, is_active
            FROM article_mapping
            WHERE source_id = %s
            ORDER BY mapping_id DESC
            """,
            (source_id,),
        )
    else:
        cur.execute(
            """
            SELECT mapping_id, source_id, source_article_id, source_article_name,
                   article_id, comment, is_active
            FROM article_mapping
            ORDER BY mapping_id DESC
            """
        )

    rows = cur.fetchall()
    cur.close()
    conn.close()

    return [
        {
            "mapping_id": r[0],
            "source_id": r[1],
            "source_article_id": r[2],
            "source_article_name": r[3],
            "article_id": r[4],
            "comment": r[5],
            "is_active": r[6],
        }
        for r in rows
    ]


@router.post("/article-mapping")
def create_article_mapping(
    source_id: int = Form(...),
    source_article_id: str = Form(""),
    source_article_name: str = Form(""),
    article_id: int = Form(...),
    comment: str = Form(""),
):
    conn = get_connection()
    cur = conn.cursor()

    # Resolve source_system from import_sources
    cur.execute("SELECT source_name FROM import_sources WHERE id = %s", (source_id,))
    row = cur.fetchone()
    source_system = row[0] if row else str(source_id)

    if source_article_id:
        cur.execute(
            """
            SELECT mapping_id FROM article_mapping
            WHERE source_id = %s AND source_article_id = %s AND is_active = TRUE
            """,
            (source_id, source_article_id),
        )
        if cur.fetchone():
            cur.close()
            conn.close()
            return {"status": "exists", "message": "Відповідність вже існує"}

    cur.execute(
        """
        INSERT INTO article_mapping
            (source_id, source_system, source_article_id, source_article_name,
             article_id, comment, is_active)
        VALUES (%s, %s, %s, %s, %s, %s, TRUE)
        RETURNING mapping_id
        """,
        (source_id, source_system, source_article_id, source_article_name,
         article_id, comment),
    )
    new_id = cur.fetchone()[0]
    conn.commit()
    cur.close()
    conn.close()

    return {"status": "ok", "mapping_id": new_id}


@router.put("/article-mapping/{mapping_id}")
def update_article_mapping(
    mapping_id: int,
    source_id: int = Form(...),
    source_article_id: str = Form(""),
    source_article_name: str = Form(""),
    article_id: int = Form(...),
    comment: str = Form(""),
    is_active: str = Form("true"),
):
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("SELECT source_name FROM import_sources WHERE id = %s", (source_id,))
    row = cur.fetchone()
    source_system = row[0] if row else str(source_id)

    cur.execute(
        """
        UPDATE article_mapping
        SET source_id = %s, source_system = %s, source_article_id = %s,
            source_article_name = %s, article_id = %s, comment = %s, is_active = %s
        WHERE mapping_id = %s
        """,
        (
            source_id, source_system, source_article_id, source_article_name,
            article_id, comment, is_active.lower() == "true", mapping_id,
        ),
    )
    conn.commit()
    cur.close()
    conn.close()

    return {"status": "ok"}


@router.delete("/article-mapping/{mapping_id}")
def delete_article_mapping(mapping_id: int):
    conn = get_connection()
    cur = conn.cursor()

    cur.execute(
        "UPDATE article_mapping SET is_active = FALSE WHERE mapping_id = %s",
        (mapping_id,),
    )
    conn.commit()
    cur.close()
    conn.close()

    return {"status": "ok"}


# ─── department mapping CRUD ──────────────────────────────────────────────────

@router.get("/department-mapping")
def get_department_mappings(source_id: Optional[int] = None):
    conn = get_connection()
    cur = conn.cursor()

    if source_id is not None:
        cur.execute(
            """
            SELECT id, source_id, external_department_code, external_department_name,
                   internal_department_id, is_active
            FROM department_import_mapping
            WHERE source_id = %s
            ORDER BY id DESC
            """,
            (source_id,),
        )
    else:
        cur.execute(
            """
            SELECT id, source_id, external_department_code, external_department_name,
                   internal_department_id, is_active
            FROM department_import_mapping
            ORDER BY id DESC
            """
        )

    rows = cur.fetchall()
    cur.close()
    conn.close()

    return [
        {
            "id": r[0],
            "source_id": r[1],
            "external_department_code": r[2],
            "external_department_name": r[3],
            "internal_department_id": r[4],
            "is_active": r[5],
        }
        for r in rows
    ]


@router.post("/department-mapping")
def create_department_mapping(
    source_id: int = Form(...),
    external_department_code: str = Form(""),
    external_department_name: str = Form(""),
    internal_department_id: str = Form(...),
):
    conn = get_connection()
    cur = conn.cursor()

    if external_department_code:
        cur.execute(
            """
            SELECT id FROM department_import_mapping
            WHERE source_id = %s AND external_department_code = %s AND is_active = TRUE
            """,
            (source_id, external_department_code),
        )
        if cur.fetchone():
            cur.close()
            conn.close()
            return {"status": "exists", "message": "Відповідність вже існує"}

    cur.execute(
        """
        INSERT INTO department_import_mapping
            (source_id, external_department_code, external_department_name,
             internal_department_id, is_active)
        VALUES (%s, %s, %s, %s, TRUE)
        RETURNING id
        """,
        (source_id, external_department_code, external_department_name, internal_department_id),
    )
    new_id = cur.fetchone()[0]
    conn.commit()
    cur.close()
    conn.close()

    return {"status": "ok", "id": new_id}


@router.put("/department-mapping/{mapping_id}")
def update_department_mapping(
    mapping_id: int,
    source_id: int = Form(...),
    external_department_code: str = Form(""),
    external_department_name: str = Form(""),
    internal_department_id: str = Form(...),
    is_active: str = Form("true"),
):
    conn = get_connection()
    cur = conn.cursor()

    cur.execute(
        """
        UPDATE department_import_mapping
        SET source_id = %s, external_department_code = %s,
            external_department_name = %s, internal_department_id = %s, is_active = %s
        WHERE id = %s
        """,
        (
            source_id, external_department_code, external_department_name,
            internal_department_id, is_active.lower() == "true", mapping_id,
        ),
    )
    conn.commit()
    cur.close()
    conn.close()

    return {"status": "ok"}


@router.delete("/department-mapping/{mapping_id}")
def delete_department_mapping(mapping_id: int):
    conn = get_connection()
    cur = conn.cursor()

    cur.execute(
        "UPDATE department_import_mapping SET is_active = FALSE WHERE id = %s",
        (mapping_id,),
    )
    conn.commit()
    cur.close()
    conn.close()

    return {"status": "ok"}


# ─── pnl column mapping CRUD ─────────────────────────────────────────────────

@router.get("/column-mapping/{source_id}")
def get_pnl_column_mapping(source_id: int):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT period_col, dept_code_col, dept_name_col,
               article_code_col, article_name_col, amount_col, comment_col
        FROM pnl_import_column_mapping
        WHERE source_id = %s
        """,
        (source_id,),
    )
    row = cur.fetchone()
    cur.close()
    conn.close()
    if not row:
        return None
    return {
        "source_id": source_id,
        "period_col": row[0] or "",
        "dept_code_col": row[1] or "",
        "dept_name_col": row[2] or "",
        "article_code_col": row[3] or "",
        "article_name_col": row[4] or "",
        "amount_col": row[5] or "",
        "comment_col": row[6] or "",
    }


@router.post("/column-mapping")
def save_pnl_column_mapping(
    source_id: int = Form(...),
    period_col: str = Form(""),
    dept_code_col: str = Form(""),
    dept_name_col: str = Form(""),
    article_code_col: str = Form(""),
    article_name_col: str = Form(""),
    amount_col: str = Form(""),
    comment_col: str = Form(""),
):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute(
        """
        INSERT INTO pnl_import_column_mapping
            (source_id, period_col, dept_code_col, dept_name_col,
             article_code_col, article_name_col, amount_col, comment_col, updated_at)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,NOW())
        ON CONFLICT (source_id) DO UPDATE SET
            period_col       = EXCLUDED.period_col,
            dept_code_col    = EXCLUDED.dept_code_col,
            dept_name_col    = EXCLUDED.dept_name_col,
            article_code_col = EXCLUDED.article_code_col,
            article_name_col = EXCLUDED.article_name_col,
            amount_col       = EXCLUDED.amount_col,
            comment_col      = EXCLUDED.comment_col,
            updated_at       = NOW()
        """,
        (source_id, period_col, dept_code_col, dept_name_col,
         article_code_col, article_name_col, amount_col, comment_col),
    )
    conn.commit()
    cur.close()
    conn.close()
    return {"status": "ok"}


# ─── preview ──────────────────────────────────────────────────────────────────

@router.post("/preview")
async def preview_source(
    source_type: str = Form(...),
    source_id: Optional[int] = Form(None),
    file: Optional[UploadFile] = File(None),
    sheet_url: Optional[str] = Form(None),
    sheet_name: Optional[str] = Form(None),
):
    try:
        if source_type == "olap_ssas_dax":
            if not source_id:
                raise HTTPException(status_code=400, detail="source_id потрібен для SSAS DAX джерела")
            rows, sheet_names = _read_ssas_dax(source_id)
        elif source_type == "sql_odbc":
            if not source_id:
                raise HTTPException(status_code=400, detail="source_id потрібен для SQL ODBC джерела")
            rows, sheet_names = _read_sql_odbc(source_id)
        elif source_type == "olap_sql":
            if not source_id:
                raise HTTPException(status_code=400, detail="source_id потрібен для OLAP джерела")
            rows, sheet_names = _read_olap(source_id)
        elif source_type == "google_sheets":
            if not sheet_url:
                raise HTTPException(
                    status_code=400, detail="Потрібно вказати посилання на Google Sheet"
                )
            rows, sheet_names = _read_google_sheet(sheet_url, sheet_name or None)
        elif source_type == "file":
            if not file:
                raise HTTPException(status_code=400, detail="Потрібно завантажити файл")
            file_bytes = await file.read()
            rows, sheet_names = _read_file(file_bytes, file.filename, sheet_name or None)
        else:
            raise HTTPException(status_code=400, detail=f"Невідомий тип джерела: {source_type}")
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))

    if not rows:
        return {"columns": [], "preview_rows": [], "total_rows": 0, "sheet_names": sheet_names}

    return {
        "columns": list(rows[0].keys()),
        "preview_rows": rows[:5],
        "total_rows": len(rows),
        "sheet_names": sheet_names,
    }


# ─── shared source reader ────────────────────────────────────────────────────

_DB_SOURCE_TYPES = {"olap_ssas_dax", "sql_odbc", "olap_sql"}


async def _read_source(source_type, file, sheet_url, sheet_name, source_id=None):
    if source_type == "olap_ssas_dax":
        if not source_id:
            raise HTTPException(status_code=400, detail="source_id потрібен для SSAS DAX джерела")
        return _read_ssas_dax(source_id)
    if source_type == "sql_odbc":
        if not source_id:
            raise HTTPException(status_code=400, detail="source_id потрібен для SQL ODBC джерела")
        return _read_sql_odbc(source_id)
    if source_type == "olap_sql":
        if not source_id:
            raise HTTPException(status_code=400, detail="source_id потрібен для OLAP джерела")
        return _read_olap(source_id)
    if source_type == "google_sheets":
        if not sheet_url:
            raise HTTPException(status_code=400, detail="Потрібно вказати посилання на Google Sheet")
        return _read_google_sheet(sheet_url, sheet_name or None)
    if source_type == "file":
        if not file:
            raise HTTPException(status_code=400, detail="Потрібно завантажити файл")
        file_bytes = await file.read()
        return _read_file(file_bytes, file.filename, sheet_name or None)
    raise HTTPException(status_code=400, detail=f"Невідомий тип джерела: {source_type}")


# ─── validate import ─────────────────────────────────────────────────────────

@router.post("/validate")
async def validate_import(
    source_id: int = Form(...),
    import_type: str = Form(...),
    scenario: str = Form(""),
    version_name: str = Form(""),
    source_type: str = Form(...),
    file: Optional[UploadFile] = File(None),
    sheet_url: Optional[str] = Form(None),
    sheet_name: Optional[str] = Form(None),
    period_col: str = Form(...),
    dept_code_col: str = Form(""),
    dept_name_col: str = Form(""),
    article_code_col: str = Form(""),
    article_name_col: str = Form(""),
    amount_col: str = Form(...),
    comment_col: str = Form(""),
):
    try:
        rows, _ = await _read_source(source_type, file, sheet_url, sheet_name, source_id)
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Помилка читання джерела: {exc}")

    if not rows:
        return {"status": "ok", "candidates": 0, "skipped": 0, "total_rows": 0,
                "errors": [], "existing_count": 0}

    conn = get_connection()
    cur = conn.cursor()
    try:
        candidates, skipped, errors = _process_rows(
            cur, rows, source_id,
            dept_code_col, dept_name_col, article_code_col, article_name_col,
            period_col, amount_col, comment_col,
        )
        periods = list({c["period"] for c in candidates})
        existing_count = _count_existing(cur, import_type, scenario, version_name, periods)
    finally:
        cur.close()
        conn.close()

    return {
        "status": "ok",
        "candidates": len(candidates),
        "skipped": skipped,
        "total_rows": len(rows),
        "errors": errors[:200],
        "existing_count": existing_count,
    }


# ─── commit import ────────────────────────────────────────────────────────────

@router.post("/commit")
async def commit_import(
    source_id: int = Form(...),
    import_type: str = Form(...),
    scenario: str = Form(""),
    version_name: str = Form(""),
    source_type: str = Form(...),
    file: Optional[UploadFile] = File(None),
    sheet_url: Optional[str] = Form(None),
    sheet_name: Optional[str] = Form(None),
    period_col: str = Form(...),
    dept_code_col: str = Form(""),
    dept_name_col: str = Form(""),
    article_code_col: str = Form(""),
    article_name_col: str = Form(""),
    amount_col: str = Form(...),
    comment_col: str = Form(""),
    replace_existing: str = Form("false"),
):
    try:
        rows, _ = await _read_source(source_type, file, sheet_url, sheet_name, source_id)
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Помилка читання джерела: {exc}")

    if not rows:
        return {"status": "ok", "imported": 0, "skipped": 0, "total_rows": 0, "errors": []}

    conn = get_connection()
    cur = conn.cursor()
    try:
        candidates, skipped, errors = _process_rows(
            cur, rows, source_id,
            dept_code_col, dept_name_col, article_code_col, article_name_col,
            period_col, amount_col, comment_col,
        )

        if replace_existing.lower() == "true" and candidates:
            periods = list({c["period"] for c in candidates})
            if import_type == "plan":
                cur.execute(
                    "DELETE FROM plan_pnl WHERE scenario = %s AND version_name = %s AND period = ANY(%s::date[])",
                    (scenario, version_name, periods),
                )
            else:
                cur.execute("DELETE FROM fact_pnl WHERE period = ANY(%s::date[])", (periods,))

        imported = 0
        for c in candidates:
            if import_type == "plan":
                cur.execute(
                    """
                    INSERT INTO plan_pnl
                        (period, holding_name, organization_name, region_name, branch_name,
                         department_id, department_name, article_id, article_name, pnl_id,
                         scenario, version_name, amount, comment, created_at, updated_at)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW(),NOW())
                    """,
                    (
                        c["period"], c["holding_name"], c["organization_name"],
                        c["region_name"], c["branch_name"],
                        c["dept_id"], c["department_name"],
                        c["art_id"], c["article_name"], c["pnl_id"],
                        scenario, version_name, c["amount"], c["comment"],
                    ),
                )
            else:
                cur.execute(
                    """
                    INSERT INTO fact_pnl
                        (period, holding_name, organization_name, region_name, branch_name,
                         department_id, department_name, article_id, article_name, pnl_id,
                         amount, registrar, source_name, loaded_at)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW())
                    """,
                    (
                        c["period"], c["holding_name"], c["organization_name"],
                        c["region_name"], c["branch_name"],
                        c["dept_id"], c["department_name"],
                        c["art_id"], c["article_name"], c["pnl_id"],
                        c["amount"], "", "",
                    ),
                )
            imported += 1

        conn.commit()

    except Exception as exc:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Помилка імпорту: {exc}")
    finally:
        cur.close()
        conn.close()

    return {
        "status": "ok",
        "imported": imported,
        "skipped": skipped,
        "total_rows": len(rows),
        "errors": errors[:100],
    }


# kept for backward compatibility
@router.post("/run")
async def run_import(
    source_id: int = Form(...),
    import_type: str = Form(...),
    scenario: str = Form(""),
    version_name: str = Form(""),
    source_type: str = Form(...),
    file: Optional[UploadFile] = File(None),
    sheet_url: Optional[str] = Form(None),
    sheet_name: Optional[str] = Form(None),
    period_col: str = Form(...),
    dept_code_col: str = Form(""),
    dept_name_col: str = Form(""),
    article_code_col: str = Form(""),
    article_name_col: str = Form(""),
    amount_col: str = Form(...),
    comment_col: str = Form(""),
    replace_existing: str = Form("true"),
):
    return await commit_import(
        source_id=source_id, import_type=import_type, scenario=scenario,
        version_name=version_name, source_type=source_type,
        file=file, sheet_url=sheet_url, sheet_name=sheet_name,
        period_col=period_col, dept_code_col=dept_code_col, dept_name_col=dept_name_col,
        article_code_col=article_code_col, article_name_col=article_name_col,
        amount_col=amount_col, comment_col=comment_col,
        replace_existing=replace_existing,
    )
